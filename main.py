from __future__ import annotations

import os, glob
import sys
from typing import Union, Dict
from types import SimpleNamespace
import random
import datetime
import shutil
from pathlib import Path

import hydra
from omegaconf import DictConfig
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter

from utils.config_utils import read_conf
from utils.cad_drawer import model_drawer
from utils.abq_connector import connector_console
from utils.abq_solving_utils import run_solver, parce_results, process_results

config_name = 'config_ss'
globalPath = str(Path.cwd())
round_decimals = 4
random.seed(1)

now = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-').split('.')[0]
now = now[:-3]


def _get_random_value(low: float, top: float) -> float:
    return np.round(random.uniform(low, top), round_decimals)

def _print_parameters(params):
    for key in params.keys():
        val = params.get(key)
        print(f'{key:10s} > {val:7f}')

def configure_xlsx(
        solver_cfg: Union[SimpleNamespace, Dict] = None,
        folder_path: str = None
       ):
    # имена и путь xls файлов; базовое имя для инпутов и .odb
    outFileNameGeom = folder_path + '/logs_' + str(solver_cfg.job_name_prefix) + '_' + str(now) + '.xlsx'
    outFileNameResult = folder_path + '/results_' + str(solver_cfg.job_name_prefix) + '_' + str(now) + '.xlsx'
    if not glob.glob(os.path.join(folder_path, outFileNameGeom)):
        # подготовка таблиц
        colNamesRes_short = pd.DataFrame(
            {
                'diameter': [],
                'h1': [],
                'h2': [],
                'h3': [],
                'h2_3rd_layer': [],
                'width_low_cut': [],
                'cell_height_1st_layer': [],
                'repeat': [],
                'fillet_a': [],
                'fillet_b': [],
                'fillet_c': [],
                'assymetry_1st_layer': [],
                'padding': [],
                'arc_offset': [],
                'thk': [],
                f"S_mises_{solver_cfg.outputs.frame_time_for_metric[0]}":[],
                f"RF_{solver_cfg.outputs.frame_time_for_metric[0]}":[],
                f"Diameter_{solver_cfg.outputs.frame_time_for_metric[0]}": [],
                f'S_mises_{solver_cfg.outputs.frame_time_for_metric[1]}': [],
                f'RF_{solver_cfg.outputs.frame_time_for_metric[1]}': [],
                f'Diameter_{solver_cfg.outputs.frame_time_for_metric[1]}': [],
                'last time': [],
                f'S_mises_last': [],
                f'RF_last': [],
                f'Diameter_last': [],
                f'Time per design': [],
                f'FEA time': [],
            })

        writerRes = pd.ExcelWriter(str(outFileNameResult), engine='xlsxwriter')
        colNamesRes_short.to_excel(writerRes, sheet_name='short', index=False)
        writerRes._save()

        # colNamesGeoms = pd.DataFrame(
        #     {'fileName': [], 'ID': [], 'HGT': [], 'Lstr': [], 'SEC': [], 'DIA': [], 'THK': [],
        #      'ANG': [], 'Lift': [], 'CVT': [], 'LAS': [], 'EM': [],
        #      'Tangent behavior': [], 'Normal Behavior': [], 'Frames': [], 'Message': [], 'Exec time': []})
        #
        # writerGeom = pd.ExcelWriter(str(outFileNameGeom), engine='xlsxwriter')
        # colNamesGeoms.to_excel(writerGeom, sheet_name='log', index=False)
        # writerGeom._save()

        wbResults = load_workbook(filename=outFileNameResult)
        # wbLog = load_workbook(filename=outFileNameGeom)
        sheet_short = wbResults['short']
        # sheet_desc = wbResults['descriptive']
    return wbResults,sheet_short,outFileNameResult

@hydra.main(config_path="config", config_name=config_name, version_base=None)
def main(cfg: DictConfig):

    # reading configuration file
    parameters, objectives, geometry_cfg, material_model, material_cfg, solver_cfg = read_conf(cfg, globalPath)

    wbResults,  sheet_short, outFileNameResult = configure_xlsx(solver_cfg,
                                                               os.path.join(globalPath,solver_cfg.results_root)
                                                               )

    first_done = False
    attempts_done = 0
    # while not first_done:
    for _idx in range(10000):
        print(f'******** currently: {_idx}')
        t_begin = datetime.datetime.now()
        if os.path.exists(os.path.join(globalPath, solver_cfg.work_root, solver_cfg.results_root)):
            shutil.rmtree(os.path.join(globalPath, solver_cfg.work_root, solver_cfg.results_root))
        # prepare set of geometric values
        curr_geometry_cfg = geometry_cfg.__dict__.copy()
        for key in curr_geometry_cfg.keys():
            if key in parameters:
                val_range = curr_geometry_cfg.get(key)
                if len(val_range) != 2:
                    print(f'Exception: error in .yaml configuration. '
                          f'Length of geometry.{key} = {len(val_range)} > {val_range}! '
                          f'Change it to [a, b]')
                    sys.exit(1)
                elif len(val_range) == 2:
                    curr_geometry_cfg.__setitem__(key, _get_random_value(low=val_range[0], top=val_range[1]))
                else:
                    curr_geometry_cfg.__setitem__(key, val_range)

        try:
            # compile step file of stent
            length = model_drawer(curr_geometry_cfg, solver_cfg.job_name_prefix)
        except Exception as e:
            print(f'\rException in model_drawer..... It`s already {attempts_done} attempt in row', end='', flush=True)
            attempts_done += 1
            continue

        for file in glob.glob(f'./{solver_cfg.work_root}/abaqus*'):
            os.remove(path=file)
        #configure .cae and inp
        connector_console(curr_geometry_cfg, length,
                          material_model, material_cfg, solver_cfg, solver_cfg.work_root,
                          'abaqus',
                          # os.path.join(globalPath,'utils','abq_cae_compiler_explicit.py'),
                          os.path.join(globalPath,'utils','abq_cae_compiler_standard_small_part.py'),
                          os.path.join(globalPath, solver_cfg.work_root,'config.json'),
                          os.getcwd())

        # _print_parameters(curr_geometry_cfg)

        t0 = datetime.datetime.now()
        message, last_frame_time = run_solver(solver_cfg, solver_cfg.work_root, 'abaqus', globalPath)
        fea_time = datetime.datetime.now() - t0
        print(f'[solver] get message: {message}. Last frame step: {last_frame_time}. '
              f'Costed time: {fea_time}')

        parce_results(solver_cfg,'abaqus', os.path.join(globalPath, solver_cfg.work_root,'config.json'))

        try:
            process_results(
                geometry_cfg=curr_geometry_cfg,
                solver_cfg=solver_cfg,
                work_path=os.path.join(globalPath, solver_cfg.work_root),
                wbResults=wbResults,
                filename=outFileNameResult,
                sheet_short=sheet_short,
                begining_time=t_begin,
                fea_time=fea_time
            )
        except:
            pass
            first_done = True

if __name__ == "__main__":
    main()

